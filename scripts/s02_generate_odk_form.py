"""
AGEVAL Step 2 — ODK XLS Form Generator
Run: python scripts/s02_generate_odk_form.py

Reads variables_personalized.csv and produces a valid ODK XLSForm
ready to upload to KoboCollect or ODK Central.
"""

import pandas as pd
import os
import yaml
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime

# ── Paths ──────────────────────────────────────────────────────────────────────
BASE   = os.path.dirname(os.path.abspath(__file__))
ROOT   = os.path.join(BASE, "..")
CFG    = os.path.join(ROOT, "config", "config.yaml")
DICT   = os.path.join(ROOT, "dictionary", "variables_personalized.csv")
OUTDIR = os.path.join(ROOT, "forms")

def load_config():
    with open(CFG) as f:
        return yaml.safe_load(f)

def load_dict():
    df = pd.read_csv(DICT)
    return df.copy()

# ── Build survey sheet rows ────────────────────────────────────────────────────
def build_survey(df, cfg):
    lang = cfg["project"]["language"]
    label_col = "label_spanish" if lang == "spanish" else "label_english"
    rows = []

    # Form metadata - beginning
    rows.append({"type": "start", "name": "start", "label::Spanish (es)": "", "label::English (en)": ""})
    rows.append({"type": "end",   "name": "end",   "label::Spanish (es)": "", "label::English (en)": ""})
    rows.append({"type": "date", "name": "surveyDate", "label::Spanish (es)": "Fecha de la encuesta", "label::English (en)": "Survey date"})
    rows.append({"type": "deviceid", "name": "deviceid", "label::Spanish (es)": "", "label::English (en)": ""})

    # Sections
    topics_order = ["quality_meta", "household", "farm", "production", "market", "finance", "inputs", "technology","geospatial"]
    section_labels = {
        "quality_meta":  ("Información del Encuestador", "Enumerator Information"),
        "household":     ("Información del Hogar",       "Household Information"),
        "farm":          ("Información de la Finca",      "Farm Information"),
        "production":    ("Producción de Café",           "Coffee Production"),
        "market":        ("Mercado y Ventas",              "Market & Sales"),
        "finance":       ("Finanzas y Crédito",            "Finance & Credit"),
        "inputs":        ("Insumos",                       "Inputs"),
        "technology":    ("Tecnología",                    "Technology Adoption"),
        "geospatial":    ("Ubicación GPS",                 "GPS Location"),
    }

    # Subsections
    subtopics_order = ["quality_meta", "household", "farm", "plot_loop", "production", "market", "finance", "inputs", "technology","geospatial"]
    subsection_labels = {
        "quality_meta":  ("Información del Encuestador", "Enumerator Information"),
        "household":     ("Información del Hogar",       "Household Information"),
        "farm":          ("Información de la Finca",      "Farm Information"),
        "plot_loop":     ("Parcelas",                     "Plots"),
        "production":    ("Producción de Café",           "Coffee Production"),
        "market":        ("Mercado y Ventas",              "Market & Sales"),
        "finance":       ("Finanzas y Crédito",            "Finance & Credit"),
        "inputs":        ("Insumos",                       "Inputs"),
        "technology":    ("Tecnología",                    "Technology Adoption"),
        "geospatial":    ("Ubicación GPS",                 "GPS Location"),
    }

    # Loop through sections
    for topic in topics_order:
        topic_rows = df[df["topic"] == topic]
        if topic_rows.empty:
            continue

        # Begin group
        label_es, label_en = section_labels.get(topic, (topic.upper(), topic.upper()))
        rows.append({
            "type":               f"begin_group",
            "name":               f"section_{topic}",
            "label::Spanish (es)": label_es,
            "label::English (en)": label_en,
            "appearance":          "field-list",
        })

        # Loop through subsections
        for subtopic in subtopics_order:
            subtopic_rows = topic_rows[topic_rows["subtopic"] == subtopic]
            if subtopic_rows.empty:
                continue
            # Begin subgroup
            label_es, label_en = subsection_labels.get(subtopic, (subtopic.upper(), subtopic.upper()))
            rows.append({
                "type":               f"begin_group",
                "name":               f"subsection_{subtopic}",
                "label::Spanish (es)": label_es,
                "label::English (en)": label_en,
                "appearance":          "field-list",
            })

            # Repeat
            current_repeat_value = None
            repeat_counter = 0  # to make repeat names unique per subtopic if needed
            def _close_repeat_if_open(rows, current_repeat_value, repeat_name):
                if current_repeat_value is not None:
                    rows.append({"type": "end_repeat", "name": repeat_name})
            active_repeat_name = None

            for _, row in subtopic_rows.iterrows():
                vname = row["variable_name"]
                qtype = row["surv_type"]

                # Repeat logic
                repeat_value = row.get("surv_repeat_count")
                has_repeat   = pd.notna(repeat_value) and str(repeat_value).strip() != ""
                repeat_value = str(repeat_value).strip() if has_repeat else None
                # Handle repeat open/close transitions
                if repeat_value != current_repeat_value:
                    # Close a previously open repeat
                    if current_repeat_value is not None:
                        rows.append({"type": "end_repeat", "name": active_repeat_name})
                        active_repeat_name = None
                    # Open a new repeat if this row starts one
                    if has_repeat:
                        repeat_counter += 1
                        active_repeat_name = f"repeat_{subtopic}_{repeat_counter}"
                        rows.append({
                            "type":                "begin_repeat",
                            "name":                active_repeat_name,
                            "label::Spanish (es)": subsection_labels.get(subtopic, (subtopic.upper(), subtopic.upper()))[0],
                            "label::English (en)": subsection_labels.get(subtopic, (subtopic.upper(), subtopic.upper()))[1],
                            "repeat_count":        repeat_value,
                        })
                    current_repeat_value = repeat_value

                # select_one / select_multiple get a list name
                if qtype in ("select_one", "select_multiple"):
                    qtype_full = f"{qtype} {vname}_choices"
                else:
                    qtype_full = qtype

                survey_row = {
                    "type":                qtype_full,
                    "name":                vname,
                    "label::Spanish (es)": row.get("label_spanish", vname),
                    "label::English (en)": row.get("label_english", vname),
                    "required":            "TRUE" if row.get("surv_required", 0) == 1 else "FALSE",
                }

                # Relevant (appear if)
                if pd.notna(row.get("surv_relevant")) and str(row["surv_relevant"]).strip():
                    survey_row["relevant"] = row["surv_relevant"]

                # Constraint
                if pd.notna(row.get("surv_constraint")) and str(row["surv_constraint"]).strip():
                    survey_row["constraint"]         = row["surv_constraint"]
                    survey_row["constraint_message"] = row.get("surv_constraint_message", "Valor inválido")
                
                rows.append(survey_row)

                # Add ODK calculate immediately after raw variable
                if row.get("surv_calculation_include", 0) == 1:
                    expr   = row.get("surv_calculation", "")
                    output = row.get("surv_calculation_output", f"{vname}_calc")
                    if pd.notna(expr) and str(expr).strip():
                        rows.append({
                            "type":                "calculate",
                            "name":                output,
                            "label::Spanish (es)": f"[calc] {output}",
                            "label::English (en)": f"[calc] {output}",
                            "calculation":         expr,
                        })

            # Close any repeat still open at the end of the subsection
            if current_repeat_value is not None:
                rows.append({"type": "end_repeat", "name": active_repeat_name})

            rows.append({"type": "end_group", "name": f"subsection_{subtopic}"})

        rows.append({"type": "end_group", "name": f"section_{topic}"})

    # Form metadata - ending
    rows.append({"type": "text", "name": "observations", "label::Spanish (es)": "Observaciones", "label::English (en)": "Observations"})

    return pd.DataFrame(rows)

# ── Build choices sheet ────────────────────────────────────────────────────────
def build_choices(df):
    rows = []
    choice_rows = df[df["surv_type"].isin(["select_one", "select_multiple"])]
    for _, row in choice_rows.iterrows():
        list_name = f"{row['variable_name']}_choices"
        choices_raw = row.get("surv_choices", "")
        if not pd.notna(choices_raw) or not str(choices_raw).strip():
            continue
        for pair in str(choices_raw).split("|"):
            pair = pair.strip()
            if ":" not in pair:
                continue
            val, lbl = pair.split(":", 1)
            rows.append({
                "list_name": list_name,
                "name":      val.strip(),
                "label":     lbl.strip(),
            })
    return pd.DataFrame(rows)

# ── Build settings sheet ───────────────────────────────────────────────────────
def build_settings(cfg):
    return pd.DataFrame([{
        "form_title":            cfg["project"]["name"],
        "form_id":               cfg["project"]["form_id"],
        "version":               cfg["project"]["form_version"],
        "default_language":      cfg["odk"]["default_language"],
    }])

# ── Style helpers ──────────────────────────────────────────────────────────────
HEADER_FILL     = PatternFill("solid", fgColor="1B5A24")
SECTION_FILL    = PatternFill("solid", fgColor="73BE7B")
SUBSECTION_FILL = PatternFill("solid", fgColor="C7E8CB")
REPEAT_FILL     = PatternFill("solid", fgColor="BCD6EE") #8DB4DE BCD6EE E3EFF9
CALC_FILL       = PatternFill("solid", fgColor="E3EFF9")

def style_sheet(ws):
    for cell in ws[1]:
        cell.fill  = HEADER_FILL
        cell.font  = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
    for row in ws.iter_rows(min_row=2):
        type_val = str(row[0].value or "")
        if "calculate" in type_val:
            for cell in row:
                cell.fill = CALC_FILL
        elif "repeat" in type_val: 
            for cell in row: 
                cell.fill = REPEAT_FILL
        elif "group" in type_val:
            for cell in row:
                name_val = str(row[1].value or "")
                if name_val.startswith("subsection_"):
                    cell.fill = SUBSECTION_FILL
                elif name_val.startswith("section_"):
                    cell.fill = SECTION_FILL
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

# ── Main ───────────────────────────────────────────────────────────────────────
def generate_form():
    cfg = load_config()
    df  = load_dict()

    survey   = build_survey(df, cfg)
    choices  = build_choices(df)
    settings = build_settings(cfg)

    ts       = datetime.now().strftime("%Y%m%d_%H%M")
    form_id  = cfg["project"]["form_id"]
    #out_path = os.path.join(OUTDIR, f"{form_id}_{ts}.xlsx")
    out_path = os.path.join(OUTDIR, f"{form_id}.xlsx")
    os.makedirs(OUTDIR, exist_ok=True)

    wb = Workbook()

    def write_sheet(wb, name, df, first=False):
        ws = wb.active if first else wb.create_sheet(name)
        ws.title = name
        ws.append(list(df.columns))
        for _, row in df.iterrows():
            ws.append([row.get(c, "") for c in df.columns])
        style_sheet(ws)
        return ws

    write_sheet(wb, "survey",   survey,   first=True)
    write_sheet(wb, "choices",  choices)
    write_sheet(wb, "settings", settings)

    wb.save(out_path)
    print(f"✅  ODK XLS form saved: {out_path}")
    print(f"    Survey rows:  {len(survey)}")
    print(f"    Choice rows:  {len(choices)}")
    return out_path

if __name__ == "__main__":
    print("Generating ODK XLS form...")
    generate_form()
